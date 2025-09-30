from fastapi import File, UploadFile, FastAPI, Depends, HTTPException, status, Form, BackgroundTasks
import io
from typing import Optional, List
import datetime
from jose import JWTError, jwt
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from starlette.middleware.gzip import GZipMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordRequestForm, OAuth2PasswordBearer
from sqlalchemy.orm import Session
from sqlalchemy import text
import os
import shutil
import uuid
import sys
sys.path.append(os.path.dirname(__file__))
from database import SessionLocal, engine
from models import (
    Base,
    Usuario,
    Cliente,
    Fornecedor,
    OrcamentoObra,
    Despesa,
    Contrato,
    ValorMaterial,
    ResumoMensal,
    ArquivoImportado,
    TabelaImportada,
    TesteLoja,
    TesteArCondicionado,
    GrupoUsuario,
    PermissaoSistema,
    PermissaoGrupo,
    Loja,
    LojaGrupo,
    ClienteGrupo,
)
from pydantic import BaseModel, ConfigDict
import asyncio
import os
import pandas as pd
import io
import uvicorn
import json
import hashlib
import zipfile
from pathlib import Path
from fastapi.responses import Response


# App e DB
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

Base.metadata.create_all(bind=engine)

# Garante índices importantes para desempenho
def ensure_indexes():
    try:
        with engine.connect() as conn:
            conn.execute(text("CREATE INDEX IF NOT EXISTS idx_usuarios_username ON usuarios(username)"))
            # Adiciona coluna grupo_id se não existir (SQLite permite ADD COLUMN)
            try:
                res = conn.execute(text("PRAGMA table_info(usuarios)"))
                cols = [row[1] for row in res.fetchall()]
                if "grupo_id" not in cols:
                    conn.execute(text("ALTER TABLE usuarios ADD COLUMN grupo_id INTEGER"))
            except Exception:
                pass
    except Exception:
        # Se falhar (ex.: permissão ou DB read-only), segue sem interromper o app
        pass

ensure_indexes()

# Garante permissões padrão do sistema (IDs fixos usados no frontend)
def ensure_system_permissions():
    DEFAULT_PERMISSIONS = [
        # Sistema
        {"id": 1001, "nome": "Dashboard", "categoria": "Sistema"},
        {"id": 1002, "nome": "Relatórios", "categoria": "Sistema"},
        {"id": 1003, "nome": "Análises", "categoria": "Sistema"},
        {"id": 1004, "nome": "Administração do Sistema", "categoria": "Sistema"},
        {"id": 1005, "nome": "Backup", "categoria": "Sistema"},
        # Cadastros - Usuários
        {"id": 1101, "nome": "Usuários", "categoria": "Cadastros"},
        {"id": 1102, "nome": "Usuários - Alterar", "categoria": "Cadastros"},
        {"id": 1103, "nome": "Usuários - Excluir", "categoria": "Cadastros"},
        {"id": 1104, "nome": "Usuários - Criar", "categoria": "Cadastros"},
        # Cadastros - Clientes
        {"id": 1201, "nome": "Clientes", "categoria": "Cadastros"},
        {"id": 1202, "nome": "Clientes - Alterar", "categoria": "Cadastros"},
        {"id": 1203, "nome": "Clientes - Excluir", "categoria": "Cadastros"},
        {"id": 1204, "nome": "Clientes - Criar/Importar", "categoria": "Cadastros"},
        # Cadastros - Fornecedores
        {"id": 1301, "nome": "Fornecedores", "categoria": "Cadastros"},
        {"id": 1302, "nome": "Fornecedores - Alterar", "categoria": "Cadastros"},
        {"id": 1303, "nome": "Fornecedores - Excluir", "categoria": "Cadastros"},
        {"id": 1304, "nome": "Fornecedores - Criar/Importar", "categoria": "Cadastros"},
        # Obras - Contratos
        {"id": 1401, "nome": "Contratos", "categoria": "Obras"},
        {"id": 1402, "nome": "Contratos - Alterar", "categoria": "Obras"},
        {"id": 1403, "nome": "Contratos - Excluir", "categoria": "Obras"},
        {"id": 1404, "nome": "Contratos - Criar/Importar", "categoria": "Obras"},
        # Obras - Orçamento
        {"id": 1501, "nome": "Orçamento de Obra", "categoria": "Obras"},
        {"id": 1502, "nome": "Orçamento - Alterar", "categoria": "Obras"},
        {"id": 1503, "nome": "Orçamento - Excluir", "categoria": "Obras"},
        {"id": 1504, "nome": "Orçamento - Criar/Importar", "categoria": "Obras"},
        # Financeiro - Despesas
        {"id": 1601, "nome": "Despesas", "categoria": "Financeiro"},
        {"id": 1602, "nome": "Despesas - Alterar", "categoria": "Financeiro"},
        {"id": 1603, "nome": "Despesas - Excluir", "categoria": "Financeiro"},
        {"id": 1604, "nome": "Despesas - Criar/Importar", "categoria": "Financeiro"},
        # Materiais
        {"id": 1701, "nome": "Valor Materiais", "categoria": "Materiais"},
        {"id": 1702, "nome": "Valor Materiais - Alterar", "categoria": "Materiais"},
        {"id": 1703, "nome": "Valor Materiais - Excluir", "categoria": "Materiais"},
        {"id": 1704, "nome": "Valor Materiais - Criar/Importar", "categoria": "Materiais"},
        # Relatórios
        {"id": 1801, "nome": "Resumo Mensal", "categoria": "Relatórios"},
        {"id": 1802, "nome": "Resumo Mensal - Alterar", "categoria": "Relatórios"},
        {"id": 1803, "nome": "Resumo Mensal - Excluir", "categoria": "Relatórios"},
        {"id": 1804, "nome": "Resumo Mensal - Criar/Importar", "categoria": "Relatórios"},
        # Lojas
        {"id": 1901, "nome": "Acesso a Todas as Lojas", "categoria": "Lojas"},
        {"id": 1902, "nome": "Acesso a Loja Individual", "categoria": "Lojas"},
        {"id": 1903, "nome": "Gerenciar Lojas", "categoria": "Lojas"},
    ]
    db = SessionLocal()
    try:
        for p in DEFAULT_PERMISSIONS:
            existing = db.query(PermissaoSistema).filter(PermissaoSistema.id == p["id"]).first()
            if not existing:
                db.add(PermissaoSistema(id=p["id"], nome=p["nome"], categoria=p["categoria"], ativo=True))
            else:
                # Mantém IDs fixos e atualiza nome/categoria se mudarem
                changed = False
                if existing.nome != p["nome"]:
                    existing.nome = p["nome"]; changed = True
                if existing.categoria != p["categoria"]:
                    existing.categoria = p["categoria"]; changed = True
                if existing.ativo is False:
                    existing.ativo = True; changed = True
                if changed:
                    db.add(existing)
        db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()

ensure_system_permissions()

# Garante existência do usuário admin no mesmo DB do backend
def ensure_admin_user():
    db = SessionLocal()
    try:
        admin = db.query(Usuario).filter(Usuario.username == "admin").first()
        desired_hash = hashlib.sha256("admin".encode()).hexdigest()
        allow_reset = (os.getenv("SEED_ADMIN", "1") == "1")
        if not admin:
            # Criar admin padrão para ambiente de desenvolvimento
            admin = Usuario(
                username="admin",
                hashed_password=desired_hash,
                nome="Administrador",
                email="admin@thors.com",
                nivel_acesso="Admin",
                ativo=True,
            )
            db.add(admin)
            db.commit()
        else:
            # Garantir que esteja ativo e com nível adequado; se senha diferente, alinhar para facilitar login
            changed = False
            if not admin.ativo:
                admin.ativo = True; changed = True
            # Normalizar nível de acesso
            if (admin.nivel_acesso or "").lower() not in {"admin", "willians"}:
                admin.nivel_acesso = "Admin"; changed = True
            # Se a senha não for 'admin', ajustar para permitir acesso imediato (ambiente dev)
            if allow_reset and admin.hashed_password != desired_hash:
                admin.hashed_password = desired_hash; changed = True
            if changed:
                db.add(admin)
                db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()

ensure_admin_user()

# Garante existência do usuário Willians (nível especial tratado como admin)
def ensure_willians_user():
    if os.getenv("SEED_WILLIANS", "1") != "1":
        return
    db = SessionLocal()
    try:
        user = db.query(Usuario).filter(Usuario.username == "willians").first()
        desired_hash = hashlib.sha256("willians".encode()).hexdigest()
        if not user:
            user = Usuario(
                username="willians",
                hashed_password=desired_hash,
                nome="Willians",
                email="willians@thors.com",
                nivel_acesso="willians",  # tratado como admin nas verificações
                ativo=True,
            )
            db.add(user)
            db.commit()
        else:
            changed = False
            if not user.ativo:
                user.ativo = True; changed = True
            if (user.nivel_acesso or "").lower() != "willians":
                user.nivel_acesso = "willians"; changed = True
            if user.hashed_password != desired_hash:
                user.hashed_password = desired_hash; changed = True
            if changed:
                db.add(user)
                db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()

ensure_willians_user()

app = FastAPI()

# Seed opcional de clientes e lojas (loja01..loja16) no start do app
def _ensure_min_clients(n: int = 16):
    db = SessionLocal()
    try:
        count = db.query(Cliente).count()
        to_create = max(0, n - count)
        if to_create > 0:
            start_idx = count + 1
            for i in range(start_idx, start_idx + to_create):
                nome = f"Cliente {i:02d}"
                db.add(Cliente(nome=nome))
            db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()

def ensure_lojas_users():
    """Garante usuários loja01..loja16 e grupos/vínculos correspondentes.

    Controlado por SEED_LOJAS ("1" por padrão). Em produção, defina SEED_LOJAS=0
    para evitar alterações automáticas.
    """
    if os.getenv("SEED_LOJAS", "1") != "1":
        return
    try:
        _ensure_min_clients(16)
        # Import tardio para evitar custo no tempo de import e dependências circulares
        from seed_manutencao_lojas import run as seed_lojas_run  # type: ignore
        seed_lojas_run()
    except Exception:
        # Não derruba o app se o seed falhar
        pass

ensure_lojas_users()

# Registro de início da aplicação para cálculo de uptime
APP_START_TIME_UTC = datetime.datetime.utcnow()

# --- Bootstrap opcional de dados para Postgres ---
def _bootstrap_sqlite_to_postgres_if_needed():
    """Se a aplicação estiver usando Postgres e a env BOOTSTRAP_FROM_SQLITE=1,
    tenta migrar dados de um arquivo SQLite local (gestao_obras.db) apenas uma vez.

    Usa um arquivo de marca em DATA_DIR/backups/.bootstrap_done para evitar repetição.
    """
    try:
        from database import SQLALCHEMY_DATABASE_URL as _URL  # type: ignore
    except Exception:
        return
    if not _URL.startswith("postgresql+"):
        return  # só relevante se destino for Postgres
    if os.getenv("BOOTSTRAP_FROM_SQLITE", "0") != "1":
        return
    force = os.getenv("BOOTSTRAP_FORCE", "0") == "1"
    # Determina um diretório gravável para o marker sem depender de variáveis globais
    try:
        marker_base = (
            os.getenv("DATA_DIR")
            or os.getenv("DB_DIR")
            or os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
            or "/tmp"
        )
        os.makedirs(marker_base, exist_ok=True)
    except Exception:
        marker_base = "."
    marker_path = os.path.join(marker_base, "backups", ".bootstrap_done")
    try:
        if os.path.exists(marker_path) and not force:
            print("[BOOTSTRAP] Já executado anteriormente (marker encontrado). Use BOOTSTRAP_FORCE=1 para forçar.")
            return
    except Exception:
        pass
    # Localiza arquivo SQLite de origem
    sqlite_candidates = [
        os.path.join(os.path.dirname(__file__), "gestao_obras.db"),
        os.path.join(os.path.dirname(__file__), "..", "gestao_obras.db"),
    ]
    sqlite_path = None
    for c in sqlite_candidates:
        if os.path.exists(c):
            sqlite_path = c
            break
    if not sqlite_path:
        return
    try:
        # Importa função migrate dinamicamente para evitar custo se não usado
        import importlib.util
        script_path = os.path.join(os.path.dirname(__file__), "migrate_sqlite_to_postgres.py")
        if not os.path.exists(script_path):
            return
        spec = importlib.util.spec_from_file_location("_migrate_mod", script_path)
        if not spec or not spec.loader:
            return
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)  # type: ignore

        print(f"[BOOTSTRAP] Migrando dados iniciais do SQLite para Postgres (force={force})...")
        # Se force ativo, truncar destino para cópia integral
        mod.migrate(sqlite_path, _URL, do_truncate=force)  # type: ignore

        # Marca como concluído
        try:
            os.makedirs(os.path.dirname(marker_path), exist_ok=True)
            with open(marker_path, "w", encoding="utf-8") as f:
                f.write(f"ok force={force}\n")
        except Exception:
            pass
    except Exception as e:
        print(f"[BOOTSTRAP][WARN] Falha ao migrar dados iniciais: {e}")

_bootstrap_sqlite_to_postgres_if_needed()

def _get_build_meta():
    """Coleta metadados de build a partir de variáveis de ambiente.

    Compatível com Render (RENDER_GIT_COMMIT) e variáveis genéricas.
    """
    return {
        "version": os.getenv("APP_VERSION"),
        "commit_sha": os.getenv("RENDER_GIT_COMMIT") or os.getenv("COMMIT_SHA") or os.getenv("GIT_COMMIT"),
        "build_time": os.getenv("BUILD_TIME"),  # ISO 8601 recomendado
    }

# Debug: estatísticas rápidas do banco
class StatsResponse(BaseModel):
    usuarios: int
    clientes: int
    testes_loja: int
    testes_ar: int
    contratos: int
    valor_materiais: int
    resumo_mensal: int
    generated_at: str

@app.get("/debug/stats", response_model=StatsResponse)
def debug_stats(db: Session = Depends(get_db)):
    try:
        usuarios = db.query(Usuario).count()
        clientes = db.query(Cliente).count()
        testes_loja = db.query(TesteLoja).count()
        testes_ar = db.query(TesteArCondicionado).count()
        contratos = db.query(Contrato).count()
        valor_materiais = db.query(ValorMaterial).count()
        resumo_mensal = db.query(ResumoMensal).count()
    except Exception:
        # Em caso de erro de schema, retorna -1
        usuarios = clientes = testes_loja = testes_ar = contratos = valor_materiais = resumo_mensal = -1
    return StatsResponse(
        usuarios=usuarios,
        clientes=clientes,
        testes_loja=testes_loja,
        testes_ar=testes_ar,
        contratos=contratos,
        valor_materiais=valor_materiais,
        resumo_mensal=resumo_mensal,
        generated_at=datetime.datetime.utcnow().isoformat() + "Z",
    )

# Health check super-rápido (sem tocar no banco)
@app.get("/healthz")
def healthz():
    meta = _get_build_meta()
    # Cálculo de uptime em segundos (desde o start do processo)
    uptime_seconds = (datetime.datetime.utcnow() - APP_START_TIME_UTC).total_seconds()
    # Info CORS (preenchido após carga do módulo; aqui usamos globals().get para evitar NameError)
    cors_info = {
        "allow_origins": globals().get("allow_origins", []),
        "allow_origin_regex": globals().get("allow_origin_regex", None),
        "netlify_url": globals().get("NETLIFY_ORIGIN", None),
    }
    return {
        "status": "ok",
        "version": meta.get("version"),
        "commit_sha": meta.get("commit_sha"),
        "build_time": meta.get("build_time"),
        "uptime_seconds": int(uptime_seconds),
        "cors": cors_info,
    }

# CORS
# Inclui automaticamente o domínio do Netlify se informado por env e também aceita qualquer subdomínio *.netlify.app via regex.
NETLIFY_ORIGIN = os.getenv("FRONTEND_NETLIFY_URL") or os.getenv("NETLIFY_URL") or os.getenv("NETLIFY_SITE_URL")

origins_env = os.getenv(
    "ALLOW_ORIGINS",
    ",".join([
        "http://localhost:3000",
        "http://localhost:3001",
        "http://localhost:3005",
        # Hosts de frontend em produção
        "https://gestao-frontend.onrender.com",
        "https://gestao-frontend-ttgd.onrender.com",
    ]),
)
allow_origins = [o.strip() for o in origins_env.split(",") if o.strip()]
# Se um domínio específico do Netlify foi fornecido por env, inclua-o explicitamente na lista
if NETLIFY_ORIGIN and NETLIFY_ORIGIN not in allow_origins:
    allow_origins.append(NETLIFY_ORIGIN)

# Regex padrão aceita localhost, onrender e qualquer subdomínio *.netlify.app; pode ser sobrescrita por ALLOW_ORIGIN_REGEX
allow_origin_regex = os.getenv(
    "ALLOW_ORIGIN_REGEX",
    r"https?://((localhost:(3000|3001|3005))|([a-z0-9-]+\.netlify\.app)|gestao-frontend[\w-]*\.onrender\.com)$",
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=allow_origins,
    allow_origin_regex=allow_origin_regex,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# GZip para respostas mais leves
app.add_middleware(GZipMiddleware, minimum_size=500)

"""Diretórios de dados graváveis

Preferimos usar um diretório de dados gravável (DATA_DIR) para evitar problemas
em plataformas com filesystem somente leitura para o código (ex.: Render).
"""
DATA_DIR = os.getenv("DATA_DIR") or os.getenv("DB_DIR")
# Valida o diretório informado via env; se falhar, tenta candidatos padrão
if DATA_DIR:
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
    except Exception:
        DATA_DIR = None  # força fallback
if not DATA_DIR:
    # Tenta caminhos padrão
    for d in ("/var/data", "/data", os.path.join(os.path.dirname(__file__), "data"), "/tmp"):
        try:
            os.makedirs(d, exist_ok=True)
            DATA_DIR = d
            break
        except Exception:
            continue
if not DATA_DIR:
    # Fallback: diretório do backend (pode falhar se for read-only)
    DATA_DIR = os.path.dirname(os.path.abspath(__file__))

# Configurar diretórios de upload (dentro do DATA_DIR)
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
TESTES_LOJA_DIR = os.path.join(UPLOAD_DIR, "testes-loja")
TESTES_AR_CONDICIONADO_DIR = os.path.join(UPLOAD_DIR, "testes-ar-condicionado")

# Criar diretórios se não existirem
os.makedirs(TESTES_LOJA_DIR, exist_ok=True)
os.makedirs(TESTES_AR_CONDICIONADO_DIR, exist_ok=True)

# Servir frontend (SPA) diretamente pelo backend, opcional
# Se FRONTEND_DIST_DIR apontar para o build do React, montamos na raiz
FRONTEND_DIST_DIR = os.getenv("FRONTEND_DIST_DIR")
if not FRONTEND_DIST_DIR:
    # Defaults razoáveis: caminho usado na imagem Docker e caminho local do repositório
    docker_front = "/app/frontend_build"
    local_front = os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend", "build")
    if os.path.isdir(docker_front):
        FRONTEND_DIST_DIR = docker_front
    elif os.path.isdir(local_front):
        FRONTEND_DIST_DIR = local_front

if FRONTEND_DIST_DIR and os.path.isdir(FRONTEND_DIST_DIR):
    try:
        app.mount("/", StaticFiles(directory=FRONTEND_DIST_DIR, html=True), name="frontend_spa")
    except Exception:
        # Se falhar por algum motivo (ex.: permissões), apenas segue com API
        pass

# Configurar arquivos estáticos APENAS para subrotas de mídia (evita conflito com /uploads/{entidade})
app.mount("/uploads/testes-loja", StaticFiles(directory=TESTES_LOJA_DIR), name="uploads_testes_loja")
app.mount(
    "/uploads/testes-ar-condicionado",
    StaticFiles(directory=TESTES_AR_CONDICIONADO_DIR),
    name="uploads_testes_ar",
)

# Diretório de backups
BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.abspath(os.path.join(BACKEND_DIR, ".."))
BACKUP_DIR = os.path.join(DATA_DIR, "backups")
os.makedirs(BACKUP_DIR, exist_ok=True)

# Helpers de backup
STATE_FILE = os.path.join(BACKUP_DIR, "state.json")
PROGRESS_FILE = os.path.join(BACKUP_DIR, "progress.json")
CANCEL_FILE = os.path.join(BACKUP_DIR, "cancel.flag")
EXCLUDE_DIRS = {
    ".git",
    "node_modules",
    "__pycache__",
    ".venv",
    ".cache",
    ".parcel-cache",
    "dist",
    "build",
    ".next",
    "backups",  # evita incluir os próprios backups
}
EXCLUDE_FILES_SUFFIX = {".pyc", ".pyo", ".log"}

def load_backup_state():
    try:
        if os.path.exists(STATE_FILE):
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if data.get("last_backup_at"):
                    data["last_backup_at"] = datetime.datetime.fromisoformat(data["last_backup_at"])  # type: ignore
                return data
    except Exception:
        pass
    return {"last_backup_at": None}

def save_backup_state(state: dict):
    data = state.copy()
    if isinstance(data.get("last_backup_at"), (datetime.datetime,)):
        data["last_backup_at"] = data["last_backup_at"].isoformat()
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def list_backup_files():
    files = []
    for name in os.listdir(BACKUP_DIR):
        if name.lower().endswith(".zip"):
            p = os.path.join(BACKUP_DIR, name)
            try:
                stat = os.stat(p)
                files.append({
                    "name": name,
                    "path": p,
                    "size": stat.st_size,
                    "created": datetime.datetime.fromtimestamp(stat.st_mtime).isoformat(),
                })
            except OSError:
                continue
    files.sort(key=lambda x: x["created"], reverse=True)
    return files

def should_exclude(rel_path: str) -> bool:
    parts = Path(rel_path).parts
    for part in parts:
        if part in EXCLUDE_DIRS:
            return True
    for suf in EXCLUDE_FILES_SUFFIX:
        if rel_path.endswith(suf):
            return True
    return False

def make_backup_zip() -> str:
    now = datetime.datetime.now()
    stamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    zip_name = f"backup_{stamp}.zip"
    zip_path = os.path.join(BACKUP_DIR, zip_name)
    inprogress = zip_path + ".inprogress"
    # Evitar concorrência
    if os.path.exists(inprogress):
        raise HTTPException(status_code=409, detail="Backup já em andamento")
    open(inprogress, "w").close()
    try:
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
            for root, dirs, files in os.walk(ROOT_DIR):
                rel_root = os.path.relpath(root, ROOT_DIR)
                if rel_root == ".":
                    rel_root = ""
                # Filtrar dirs in-place para acelerar
                dirs[:] = [d for d in dirs if not should_exclude(os.path.join(rel_root, d))]
                for fn in files:
                    rel_fp = os.path.join(rel_root, fn)
                    if should_exclude(rel_fp):
                        continue
                    abs_fp = os.path.join(ROOT_DIR, rel_fp)
                    # Segurança: garantir que está dentro do ROOT_DIR
                    if not os.path.abspath(abs_fp).startswith(ROOT_DIR):
                        continue
                    zf.write(abs_fp, arcname=rel_fp)
        # Atualiza estado
        state = load_backup_state()
        state["last_backup_at"] = now
        save_backup_state(state)
        return zip_name
    finally:
        try:
            os.remove(inprogress)
        except OSError:
            pass

def collect_backup_candidates() -> list[str]:
    candidates: list[str] = []
    for root, dirs, files in os.walk(ROOT_DIR):
        rel_root = os.path.relpath(root, ROOT_DIR)
        if rel_root == ".":
            rel_root = ""
        dirs[:] = [d for d in dirs if not should_exclude(os.path.join(rel_root, d))]
        for fn in files:
            rel_fp = os.path.join(rel_root, fn)
            if should_exclude(rel_fp):
                continue
            candidates.append(rel_fp)
    return candidates

def load_progress():
    try:
        if os.path.exists(PROGRESS_FILE):
            with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        return None
    return None

def save_progress(data: dict | None):
    if data is None:
        try:
            if os.path.exists(PROGRESS_FILE):
                os.remove(PROGRESS_FILE)
        except OSError:
            pass
        return
    tmp = data.copy()
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(tmp, f, ensure_ascii=False)

def set_cancel():
    try:
        open(CANCEL_FILE, "w").close()
    except Exception:
        pass

def clear_cancel():
    try:
        if os.path.exists(CANCEL_FILE):
            os.remove(CANCEL_FILE)
    except OSError:
        pass

def is_canceled() -> bool:
    return os.path.exists(CANCEL_FILE)

async def backup_worker_async():
    now = datetime.datetime.now()
    stamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    zip_name = f"backup_{stamp}.zip"
    zip_path = os.path.join(BACKUP_DIR, zip_name)
    inprogress = zip_path + ".inprogress"
    if os.path.exists(inprogress):
        raise HTTPException(status_code=409, detail="Backup já em andamento")
    open(inprogress, "w").close()
    canceled = False
    try:
        files = collect_backup_candidates()
        total = len(files)
        processed = 0
        save_progress({"running": True, "percent": 0, "processed": 0, "total": total, "current": None, "file": zip_name, "canceled": False})
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
            for rel_fp in files:
                if is_canceled():
                    canceled = True
                    break
                abs_fp = os.path.join(ROOT_DIR, rel_fp)
                if not os.path.abspath(abs_fp).startswith(ROOT_DIR):
                    continue
                try:
                    zf.write(abs_fp, arcname=rel_fp)
                except Exception:
                    # ignora arquivo com erro
                    pass
                processed += 1
                if processed % 20 == 0 or processed == total:
                    percent = 0 if total == 0 else round(processed * 100 / total, 2)
                    save_progress({"running": True, "percent": percent, "processed": processed, "total": total, "current": rel_fp, "file": zip_name, "canceled": False})
                    await asyncio.sleep(0)
        if canceled:
            # remover zip parcial
            try:
                if os.path.exists(zip_path):
                    os.remove(zip_path)
            except OSError:
                pass
        else:
            # Finaliza com sucesso
            state = load_backup_state()
            state["last_backup_at"] = now
            save_backup_state(state)
            rotate_backups(keep=7)
    finally:
        # Marca estado final (sucesso 100% ou cancelado com percent atual)
        try:
            p = load_progress() or {}
            if canceled:
                save_progress({
                    "running": False,
                    "percent": float(p.get("percent") or 0),
                    "processed": int(p.get("processed") or 0),
                    "total": int(p.get("total") or 0),
                    "current": None,
                    "file": p.get("file"),
                    "canceled": True
                })
            else:
                total = int(p.get("total") or 0)
                save_progress({
                    "running": False,
                    "percent": 100.0,
                    "processed": total,
                    "total": total,
                    "current": None,
                    "file": p.get("file"),
                    "canceled": False
                })
        except Exception:
            pass
        try:
            os.remove(inprogress)
        except OSError:
            pass
        clear_cancel()

def rotate_backups(keep: int = 7):
    files = list_backup_files()
    if len(files) <= keep:
        return
    for f in files[keep:]:
        try:
            os.remove(os.path.join(BACKUP_DIR, f["name"]))
        except OSError:
            continue

async def schedule_backup_task():
    # Tarefa que agenda backup diariamente às 22:00
    while True:
        now = datetime.datetime.now()
        target = now.replace(hour=22, minute=0, second=0, microsecond=0)
        if now >= target:
            target = target + datetime.timedelta(days=1)
        wait_seconds = (target - now).total_seconds()
        try:
            await asyncio.sleep(wait_seconds)
            make_backup_zip()
            rotate_backups(keep=7)
        except Exception:
            # Não interromper o laço por erro
            await asyncio.sleep(5)

@app.on_event("startup")
async def _startup_schedule():
    # Evitar múltiplas tarefas em reload: usar um flag global
    if not getattr(app.state, "_backup_task_started", False):
        app.state._backup_task_started = True
        asyncio.create_task(schedule_backup_task())

# Configurações de Segurança
SECRET_KEY = os.getenv("SECRET_KEY", "secret-key-thors-gestor-2025")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# --- Schemas de Autenticação ---
class UserLogin(BaseModel):
    username: str
    password: str

class UserResponse(BaseModel):
    id: int
    username: str
    nome: str
    email: Optional[str] = None
    nivel_acesso: str
    ativo: bool
    is_admin: bool

class Token(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse

# --- Funções de Autenticação ---
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return hash_password(plain_password) == hashed_password

def authenticate_user(db: Session, username: str, password: str):
    user = db.query(Usuario).filter(Usuario.username == username).first()
    if not user:
        return False
    if not verify_password(password, user.hashed_password):
        return False
    return user

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.datetime.utcnow() + datetime.timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user = db.query(Usuario).filter(Usuario.username == username).first()
    if user is None:
        raise credentials_exception
    return user

def require_admin(current_user: Usuario = Depends(get_current_user)):
    if current_user.nivel_acesso != "Admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acesso negado. Apenas administradores podem acessar esta funcionalidade."
        )
    return current_user

def require_maintenance_or_admin(current_user: Usuario = Depends(get_current_user)):
    if current_user.nivel_acesso not in ["Admin", "Manutenção"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acesso negado. Apenas usuários de Manutenção ou Administradores podem acessar esta funcionalidade."
        )
    return current_user

# --- Permissões por ação (CRUD) ---
# Convenção usada no sistema:
# base (xx01) = leitura/listagem, editar (xx02), excluir (xx03), criar/importar (xx04)
ACTION_OFFSETS = {
    "read": 0,    # xx01
    "update": 1,  # xx02
    "delete": 2,  # xx03
    "create": 3,  # xx04
}

def _collect_permission_ids_for_user(db: Session, user: Usuario) -> set[int]:
    # Admin e Willians têm acesso total
    if str(user.nivel_acesso or '').lower() in {"admin", "willians"}:
        # retornar um set grande não é necessário; o require_permission trata admin separadamente
        return set()
    if not user.grupo_id:
        return set()
    # Buscar permissões do grupo
    links = db.query(PermissaoGrupo).filter(PermissaoGrupo.grupo_id == user.grupo_id).all()
    return {lk.permissao_id for lk in links}

def _permission_required(base_id: int, action: str):
    def dependency(current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)):
        # Admin e Willians sempre passam
        if str(current_user.nivel_acesso or '').lower() in {"admin", "willians"}:
            return current_user
        ids = _collect_permission_ids_for_user(db, current_user)
        # Sempre aceitar se tiver a base (leitura) para qualquer ação de leitura; para outras, exigir o id específico
        offset = ACTION_OFFSETS.get(action, 0)
        required = base_id + offset
        # A permissão base (xx01) concede leitura
        if action == "read":
            if (base_id in ids) or (required in ids):
                return current_user
        else:
            if (required in ids):
                return current_user
        raise HTTPException(status_code=403, detail="Permissão negada: ação não autorizada")
    return dependency

# --- Schemas ---
class ClienteCreateSchema(BaseModel):
    nome: str
    cnpj: Optional[str] = None
    email: Optional[str] = None
    contato: Optional[str] = None
    endereco: Optional[str] = None
    model_config = ConfigDict(from_attributes=True)

class ClienteSchema(BaseModel):
    id: Optional[int] = None
    nome: str
    cnpj: Optional[str] = None
    email: Optional[str] = None
    contato: Optional[str] = None
    endereco: Optional[str] = None
    model_config = ConfigDict(from_attributes=True)

class FornecedorSchema(BaseModel):
    id: Optional[int]
    nome: str
    cnpj: Optional[str]
    contato: Optional[str]
    model_config = ConfigDict(from_attributes=True)

class OrcamentoObraSchema(BaseModel):
    id: Optional[int]
    cliente_id: int
    etapa: str
    descricao: str
    unidade: str
    quantidade: float
    custo_unitario: float
    data: Optional[datetime.date]
    model_config = ConfigDict(from_attributes=True)

class DespesaSchema(BaseModel):
    id: Optional[int] = None
    id_cliente: Optional[int] = None  # Novo campo principal
    cliente_id: Optional[int] = None  # Mantido para compatibilidade
    servico: Optional[str] = None  # Novo campo principal
    descricao: Optional[str] = None  # Mantido para compatibilidade
    valor: float
    data: Optional[datetime.date] = None
    categoria: Optional[str] = None  # Novo campo
    status: Optional[str] = 'Pendente'  # Novo campo
    observacoes: Optional[str] = None  # Novo campo
    model_config = ConfigDict(from_attributes=True)

class ContratoSchema(BaseModel):
    id: Optional[int] = None
    numero: str
    cliente_id: int
    cliente: Optional[str] = None  # Nome do cliente para compatibilidade
    valor: float
    dataInicio: Optional[datetime.date] = None
    dataFim: Optional[datetime.date] = None
    tipo: Optional[str] = None
    situacao: Optional[str] = None
    prazoPagamento: Optional[str] = None
    quantidadeParcelas: Optional[str] = None
    arquivo: Optional[str] = None
    arquivo_upload_id: Optional[int] = None
    model_config = ConfigDict(from_attributes=True)

class ValorMaterialSchema(BaseModel):
    id: Optional[int] = None
    cliente_id: Optional[int] = None  # Cliente específico (opcional)
    descricao_produto: str
    marca: str = ""
    unidade_medida: str = ""
    valor_unitario: float
    estoque_atual: int = 0
    estoque_minimo: int = 0
    data_ultima_entrada: str = ""
    responsavel: str = ""
    fornecedor: str = ""
    valor: Optional[float] = None
    localizacao: str = ""
    observacoes: str = ""
    model_config = ConfigDict(from_attributes=True)

class ResumoMensalSchema(BaseModel):
    id: Optional[int]
    cliente_id: int
    mes: str
    ano: int
    total_despesas: float
    total_orcamento: float
    model_config = ConfigDict(from_attributes=True)

class ArquivoImportadoSchema(BaseModel):
    id: int | None = None
    nome: str
    entidade: str
    tamanho: int
    # Alinhado com models.ArquivoImportado.criado_em (DateTime)
    # Em Pydantic v2, datetime é serializado automaticamente em ISO 8601
    criado_em: Optional[datetime.datetime] = None
    model_config = ConfigDict(from_attributes=True)

class TabelaImportadaSchema(BaseModel):
    id: int | None = None
    entidade: str
    dados: list
    upload_id: int | None = None
    criado_em: Optional[datetime.datetime] = None
    model_config = ConfigDict(from_attributes=True)

# --- Schemas para Uploads ---
class UploadResult(BaseModel):
    upload_id: int
    filename: str
    entidade: str
    records_imported: int = 0

class TesteLojaSchema(BaseModel):
    id: int | None = None
    data_teste: datetime.date
    cliente_id: int
    horario: datetime.time
    foto: str | None = None
    video: str | None = None
    status: str  # 'OK' ou 'OFF'
    observacao: str | None = None
    criado_em: Optional[datetime.datetime] = None
    model_config = ConfigDict(from_attributes=True)

class TesteLojaCreateSchema(BaseModel):
    data_teste: datetime.date
    cliente_id: int
    horario: datetime.time
    status: str
    observacao: str | None = None

class TesteArCondicionadoSchema(BaseModel):
    id: int | None = None
    data_teste: datetime.date
    cliente_id: int
    horario: datetime.time
    foto: str | None = None
    video: str | None = None
    status: str  # 'OK' ou 'OFF'
    observacao: str | None = None
    criado_em: Optional[datetime.datetime] = None
    model_config = ConfigDict(from_attributes=True)

class TesteArCondicionadoCreateSchema(BaseModel):
    data_teste: datetime.date
    cliente_id: int
    horario: datetime.time
    status: str
    observacao: str | None = None

# --- Schemas de Backup ---
class BackupStatus(BaseModel):
    last_backup_at: Optional[datetime.datetime]
    pending: bool
    hours_since: Optional[float]
    days_since: Optional[float]
    backups_count: int
    over_limit: bool
    files: list

class DeleteBackupsRequest(BaseModel):
    files: List[str]

# --- Endpoints de Backup ---
@app.get("/backup/status", response_model=BackupStatus)
def backup_status(current_user: Usuario = Depends(require_admin)):
    state = load_backup_state()
    last = state.get("last_backup_at")
    now = datetime.datetime.now()
    hours_since = None
    days_since = None
    if last:
        delta = now - last
        hours_since = round(delta.total_seconds() / 3600, 2)
        days_since = round(delta.total_seconds() / 86400, 2)
    files = list_backup_files()
    # Considera pendente se passou do horário de 22:00 do último dia sem backup
    pending = False
    if last is None:
        pending = True
    else:
        # Limiares: 22h do dia
        today_22 = now.replace(hour=22, minute=0, second=0, microsecond=0)
        if now >= today_22:
            # Depois das 22h de hoje, pendente se não houve backup desde 22h de hoje
            pending = last < today_22
        else:
            # Antes das 22h de hoje, pendente se não houve backup desde 22h de ontem
            yesterday_22 = today_22 - datetime.timedelta(days=1)
            pending = last < yesterday_22
    over_limit = len(files) > 7
    return BackupStatus(
        last_backup_at=last,
        pending=pending,
        hours_since=hours_since,
        days_since=days_since,
        backups_count=len(files),
        over_limit=over_limit,
        files=files,
    )

@app.post("/backup/run")
def run_backup(background_tasks: BackgroundTasks, current_user: Usuario = Depends(require_admin)):
    # dispara tarefa em background com progresso
    # se já houver progresso em andamento, retornar 409
    progress = load_progress()
    if progress and progress.get("running"):
        raise HTTPException(status_code=409, detail="Backup em andamento")
    # iniciar tarefa
    background_tasks.add_task(backup_worker_async)
    return {"message": "Backup iniciado"}

@app.get("/backup/list")
def list_backups(current_user: Usuario = Depends(require_admin)):
    return list_backup_files()

@app.get("/backup/download/{filename}")
def download_backup(filename: str, current_user: Usuario = Depends(require_admin)):
    safe_name = os.path.basename(filename)
    path = os.path.join(BACKUP_DIR, safe_name)
    if not os.path.exists(path) or not path.startswith(BACKUP_DIR):
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")

    def iterfile():
        with open(path, "rb") as f:
            while True:
                chunk = f.read(1024 * 1024)
                if not chunk:
                    break
                yield chunk

    return StreamingResponse(iterfile(), media_type="application/zip", headers={
        "Content-Disposition": f"attachment; filename={safe_name}"
    })

@app.delete("/backup")
def delete_backups(payload: DeleteBackupsRequest, current_user: Usuario = Depends(require_admin)):
    deleted = []
    for name in payload.files:
        safe_name = os.path.basename(name)
        path = os.path.join(BACKUP_DIR, safe_name)
        if os.path.exists(path) and path.startswith(BACKUP_DIR):
            try:
                os.remove(path)
                deleted.append(safe_name)
            except OSError:
                continue
    return {"deleted": deleted}

@app.get("/backup/progress")
def backup_progress(current_user: Usuario = Depends(require_admin)):
    p = load_progress() or {}
    running = bool(p.get("running"))
    percent = float(p.get("percent") or 0)
    processed = int(p.get("processed") or 0)
    total = int(p.get("total") or 0)
    current = p.get("current")
    file = p.get("file")
    canceled = bool(p.get("canceled") or False)
    return {"running": running, "percent": percent, "processed": processed, "total": total, "current": current, "file": file, "canceled": canceled}

@app.post("/backup/cancel")
def backup_cancel(current_user: Usuario = Depends(require_admin)):
    # sinaliza cancelamento; worker irá interromper e limpar zip parcial
    if load_progress() and load_progress().get("running"):
        set_cancel()
        return {"message": "Cancelamento solicitado"}
    return {"message": "Nenhum backup em andamento"}

# --- CRUD Endpoints ---
# Usuários
class UsuarioSchema(BaseModel):
    id: int | None = None
    username: str
    nome: str | None = None
    email: str | None = None
    nivel_acesso: str = 'visualizacao'  # admin, willians, manutencao, visualizacao
    ativo: bool = True
    grupo_id: int | None = None
    # permissoes: list[str] | None = None  # Lista de permissões específicas
    model_config = ConfigDict(from_attributes=True)

class UsuarioCreateSchema(BaseModel):
    username: str
    password: str
    nome: str
    email: str | None = None
    nivel_acesso: str = 'visualizacao'  # admin, willians, manutencao, visualizacao
    ativo: bool = True
    grupo_id: int | None = None
    # permissoes: list[str] | None = None  # Lista de permissões específicas

class UsuarioUpdateSchema(BaseModel):
    username: str
    password: str | None = None  # Opcional para updates
    nome: str
    email: str | None = None
    nivel_acesso: str = 'visualizacao'  # admin, willians, manutencao, visualizacao
    ativo: bool = True
    grupo_id: int | None = None

# Grupos de Usuários
class GrupoUsuarioSchema(BaseModel):
    id: int | None = None
    nome: str
    descricao: str | None = None
    status: str = 'Aprovado'
    motivo: str | None = None
    valor_maximo_diario_financeiro: float = 0.00
    preco_venda: float = 0.00
    plano_contas: float = 0.00
    valor_maximo_movimentacao: float = 0.00
    valor_maximo_solicitacao_compra: float = 0.00
    valor_maximo_diario_solicitacao_compra: float = 0.00
    model_config = ConfigDict(from_attributes=True)

class GrupoUsuarioCreateSchema(BaseModel):
    nome: str
    descricao: str | None = None
    status: str = 'Aprovado'
    motivo: str | None = None
    valor_maximo_diario_financeiro: float = 0.00
    preco_venda: float = 0.00
    plano_contas: float = 0.00
    valor_maximo_movimentacao: float = 0.00
    valor_maximo_solicitacao_compra: float = 0.00
    valor_maximo_diario_solicitacao_compra: float = 0.00
    permissoes: List[str] = []
    lojas: List[int] = []
    acesso_total_lojas: bool = False
    clientes: List[int] = []

# Permissões
class PermissaoSistemaSchema(BaseModel):
    id: int | None = None
    nome: str
    descricao: str | None = None
    categoria: str | None = None
    ativo: bool = True
    model_config = ConfigDict(from_attributes=True)

# Lojas
class LojaSchema(BaseModel):
    id: int | None = None
    nome: str
    endereco: str | None = None
    cidade: str | None = None
    ativa: bool = True
    model_config = ConfigDict(from_attributes=True)

class LojaCreateSchema(BaseModel):
    nome: str
    endereco: str | None = None
    cidade: str | None = None
    ativa: bool = True

# --- CRUD Endpoints ---

# Utilitários de acesso por cliente
def _get_allowed_client_ids(db: Session, user: Usuario) -> Optional[List[int]]:
    """Retorna lista de cliente_ids permitidos para o usuário via grupo.
    Retorna None para acesso total (admin/willians ou sem grupo vinculado).
    """
    nivel = (user.nivel_acesso or "").lower()
    if nivel in ["admin", "willians"]:
        return None  # acesso total
    if not user.grupo_id:
        return None  # sem grupo específico => não filtra
    links = db.query(ClienteGrupo).filter(ClienteGrupo.grupo_id == user.grupo_id).all()
    ids = [lk.cliente_id for lk in links]
    return ids

class MyClientesResponse(BaseModel):
    grupo_id: int | None
    clientes: List[int]

@app.get("/me/clientes", response_model=MyClientesResponse)
def get_my_clientes(current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)):
    ids = _get_allowed_client_ids(db, current_user)
    if ids is None:
        # None significa acesso total; por compatibilidade retornamos todos IDs
        all_ids = [c.id for c in db.query(Cliente.id).all()]
        return {"grupo_id": current_user.grupo_id, "clientes": all_ids}
    return {"grupo_id": current_user.grupo_id, "clientes": ids}

@app.get("/usuarios/", response_model=List[UsuarioSchema])
def listar_usuarios(db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1101, "read"))):
    return db.query(Usuario).all()

@app.post("/usuarios/", response_model=UsuarioSchema)
def criar_usuario(usuario: UsuarioCreateSchema, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1101, "create"))):
    if db.query(Usuario).filter(Usuario.username == usuario.username).first():
        raise HTTPException(status_code=400, detail="Usuário já existe")
    novo = Usuario(
        username=usuario.username,
        nome=usuario.nome,
        email=usuario.email,
        hashed_password=Usuario.hash_password(usuario.password),
        nivel_acesso=usuario.nivel_acesso,
        ativo=usuario.ativo,
        grupo_id=usuario.grupo_id,
    )
    db.add(novo)
    db.commit()
    db.refresh(novo)
    return novo

@app.put("/usuarios/{usuario_id}", response_model=UsuarioSchema)
def editar_usuario(usuario_id: int, usuario: UsuarioUpdateSchema, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1101, "update"))):
    db_usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not db_usuario:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    db_usuario.username = usuario.username
    db_usuario.nome = usuario.nome
    db_usuario.email = usuario.email
    # Só atualizar senha se foi fornecida
    if usuario.password:
        db_usuario.hashed_password = Usuario.hash_password(usuario.password)
    db_usuario.nivel_acesso = usuario.nivel_acesso
    db_usuario.ativo = usuario.ativo
    db_usuario.grupo_id = usuario.grupo_id
    db.commit()
    db.refresh(db_usuario)
    return db_usuario

@app.delete("/usuarios/{usuario_id}")
def deletar_usuario(usuario_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1101, "delete"))):
    db_usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not db_usuario:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    db.delete(db_usuario)
    db.commit()
    return {"ok": True}

# Auxiliares de usuários por grupo
@app.get("/grupos/{grupo_id}/usuarios", response_model=List[UsuarioSchema])
def listar_usuarios_por_grupo(grupo_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1101, "read"))):
    return db.query(Usuario).filter(Usuario.grupo_id == grupo_id).all()

class SetGrupoSchema(BaseModel):
    grupo_id: int | None = None

@app.post("/usuarios/{usuario_id}/set-grupo", response_model=UsuarioSchema)
def set_grupo_usuario(usuario_id: int, payload: SetGrupoSchema, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1101, "update"))):
    user = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    user.grupo_id = payload.grupo_id
    db.commit()
    db.refresh(user)
    return user

# Login
class LoginRequest(BaseModel):
    username: str
    password: str

@app.post("/login/", response_model=Token)
def login(request: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(Usuario).filter(Usuario.username == request.username).first()
    if not user or not user.verify_password(request.password):
        raise HTTPException(status_code=401, detail="Usuário ou senha inválidos")
    
    token_data = {
        "sub": user.username,
        "exp": datetime.datetime.utcnow() + datetime.timedelta(hours=12),
    }
    token = jwt.encode(token_data, SECRET_KEY, algorithm=ALGORITHM)
    
    # Criar resposta do usuário
    user_response = UserResponse(
        id=user.id,
        username=user.username,
        nome=user.nome,
        email=user.email,
        nivel_acesso=user.nivel_acesso,
        ativo=user.ativo,
        is_admin=(user.nivel_acesso.lower() in ["admin", "willians"])
    )
    
    return Token(
        access_token=token,
        token_type="bearer",
        user=user_response
    )

@app.get("/me/", response_model=UserResponse)
def get_current_user_info(current_user: Usuario = Depends(get_current_user)):
    """Obter informações do usuário atual"""
    return UserResponse(
        id=current_user.id,
        username=current_user.username,
        nome=current_user.nome,
        email=current_user.email,
        nivel_acesso=current_user.nivel_acesso,
        ativo=current_user.ativo,
        is_admin=(current_user.nivel_acesso.lower() in ["admin", "willians"])
    )

# --- Uploads genéricos de arquivos (Excel e outros) ---
# Nota: mantemos endpoints legacy sob "/uploads/*" MAS adicionamos aliases em "/api/uploads/*"
# para evitar conflito com StaticFiles montado em "/uploads" (usado para servir mídias).
@app.post("/uploads/{entidade}", response_model=UploadResult)
@app.post("/api/uploads/{entidade}", response_model=UploadResult)
async def upload_entidade(entidade: str, file: UploadFile = File(...), db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    # Persistir arquivo no banco (ArquivosImportados)
    content = await file.read()
    up = ArquivoImportado(
        nome=file.filename,
        entidade=entidade,
        conteudo=content,
        tamanho=len(content),
    )
    db.add(up)
    db.commit()
    db.refresh(up)

    # Opcional: se for valor_materiais ou clientes, tentar parse simples (CSV/XLSX)
    imported = 0
    if entidade == "valor_materiais" and file.filename.lower().endswith((".csv", ".xlsx", ".xls")):
        try:
            import pandas as pd  # type: ignore
            buf = io.BytesIO(content)
            if file.filename.lower().endswith('.csv'):
                df = pd.read_csv(buf)
            else:
                df = pd.read_excel(buf)
            # Normalização de colunas esperadas
            cols = {c.lower().strip(): c for c in df.columns}
            def pick(row, key, default=None):
                # aceita chave exata ou variações comuns
                candidates = [key, key.replace('_', ' '), key.replace('_', ''), key.replace('_', '-')]
                for k in candidates:
                    if k in cols:
                        return row[cols[k]]
                return default
            for _, row in df.iterrows():
                vm = ValorMaterial(
                    cliente_id=None,
                    descricao_produto=str(pick(row, 'descricao_produto', pick(row, 'descricao', '')) or ''),
                    marca=str(pick(row, 'marca', '') or ''),
                    unidade_medida=str(pick(row, 'unidade_medida', pick(row, 'unidade', '')) or ''),
                    valor_unitario=float(pick(row, 'valor_unitario', pick(row, 'valor', 0)) or 0),
                    estoque_atual=int(pick(row, 'estoque_atual', 0) or 0),
                    estoque_minimo=int(pick(row, 'estoque_minimo', 0) or 0),
                    data_ultima_entrada=str(pick(row, 'data_ultima_entrada', '') or ''),
                    responsavel=str(pick(row, 'responsavel', '') or ''),
                    fornecedor=str(pick(row, 'fornecedor', '') or ''),
                    valor=float(pick(row, 'valor', 0) or 0),
                    localizacao=str(pick(row, 'localizacao', '') or ''),
                    observacoes=str(pick(row, 'observacoes', '') or ''),
                )
                db.add(vm)
                imported += 1
            db.commit()
        except Exception:
            # Se falhar o parse, não impede o upload; apenas segue com 0 importados
            db.rollback()
            imported = 0
    elif entidade == "clientes" and file.filename.lower().endswith((".csv", ".xlsx", ".xls")):
        try:
            import pandas as pd  # type: ignore
            buf = io.BytesIO(content)
            if file.filename.lower().endswith('.csv'):
                df = pd.read_csv(buf)
            else:
                df = pd.read_excel(buf)

            # Normalização de colunas esperadas
            cols = {str(c).strip().lower(): c for c in df.columns}

            def col(row, key, default=None):
                variants = [
                    key,
                    key.replace('_', ' '),
                    key.replace('_', ''),
                    key.replace('_', '-'),
                ]
                for k in variants:
                    if k in cols:
                        return row[cols[k]]
                return default

            for _, row in df.iterrows():
                _id = col(row, 'id')
                nome = str(col(row, 'nome', '') or '').strip()
                cnpj = str(col(row, 'cnpj', '') or '').strip()
                email = str(col(row, 'email', '') or '').strip()
                contato = str(col(row, 'contato', '') or '').strip()
                endereco = str(col(row, 'endereco', col(row, 'endereço', '')) or '').strip()

                # Pular linhas sem nome
                if not nome:
                    continue

                # Estratégia de upsert: tenta localizar por ID, depois por CNPJ, depois por nome
                existing = None
                if _id not in (None, "", float('nan')):
                    try:
                        _id_int = int(_id)
                        existing = db.query(Cliente).filter(Cliente.id == _id_int).first()
                    except Exception:
                        existing = None
                if not existing and cnpj:
                    existing = db.query(Cliente).filter(Cliente.cnpj == cnpj).first()
                if not existing and nome:
                    existing = db.query(Cliente).filter(Cliente.nome == nome).first()

                if existing:
                    # Atualiza campos conhecidos
                    changed = False
                    if cnpj and existing.cnpj != cnpj:
                        existing.cnpj = cnpj; changed = True
                    if email and existing.email != email:
                        existing.email = email; changed = True
                    if contato and existing.contato != contato:
                        existing.contato = contato; changed = True
                    if endereco and existing.endereco != endereco:
                        existing.endereco = endereco; changed = True
                    if changed:
                        db.add(existing)
                        imported += 1
                else:
                    novo = Cliente(
                        nome=nome,
                        cnpj=cnpj or None,
                        email=email or None,
                        contato=contato or None,
                        endereco=endereco or None,
                    )
                    db.add(novo)
                    imported += 1
            db.commit()
        except Exception as e:
            # Em caso de erro, faz rollback e segue
            print(f"[UPLOAD clientes] Erro ao importar: {e}")
            db.rollback()
            imported = 0

    elif entidade == "contratos":
        # Importação ESTRITA via modelo: aceita somente .xlsx/.xls com colunas exatamente iguais ao template
        try:
            if not file.filename.lower().endswith((".xlsx", ".xls")):
                raise HTTPException(status_code=400, detail="Apenas arquivos Excel (.xlsx/.xls) são aceitos para contratos.")
            import pandas as pd  # type: ignore
            buf = io.BytesIO(content)
            df = pd.read_excel(buf)

            expected = [
                "numero",
                "cliente_id",
                "valor",
                "dataInicio",
                "dataFim",
                "tipo",
                "situacao",
                "prazoPagamento",
                "quantidadeParcelas",
            ]
            # Normalizar nomes (case-insensitive), mas exigir conjunto exato de colunas
            norm = [str(c).strip() for c in df.columns]
            if set(norm) != set(expected) or len(norm) != len(expected):
                raise HTTPException(
                    status_code=400,
                    detail=f"Colunas inválidas. Use exatamente estas colunas: {', '.join(expected)}",
                )

            # Helper de datas
            from datetime import date, datetime as dt

            def to_date(v):
                if pd.isna(v) or v is None or v == "":
                    return None
                if isinstance(v, (dt, date)):
                    return v.date() if isinstance(v, dt) else v
                # tentar parse a partir de string
                try:
                    return pd.to_datetime(v).date()
                except Exception:
                    return None

            errors: list[str] = []
            new_count = 0

            # Validar todas as linhas antes (transação all-or-nothing)
            rows_data = []
            for idx, row in df.iterrows():
                numero = str(row.get("numero", "")).strip()
                cliente_id = row.get("cliente_id", None)
                valor = row.get("valor", None)
                dataInicio = to_date(row.get("dataInicio", None))
                dataFim = to_date(row.get("dataFim", None))
                tipo = str(row.get("tipo", "")).strip() or None
                situacao = str(row.get("situacao", "")).strip() or None
                prazoPagamento = str(row.get("prazoPagamento", "")).strip() or None
                quantidadeParcelas = str(row.get("quantidadeParcelas", "")).strip() or None

                if not numero:
                    errors.append(f"Linha {idx+2}: campo 'numero' é obrigatório.")
                try:
                    cliente_id_int = int(cliente_id)
                except Exception:
                    errors.append(f"Linha {idx+2}: 'cliente_id' inválido.")
                    cliente_id_int = None  # type: ignore
                try:
                    valor_num = float(valor)
                except Exception:
                    errors.append(f"Linha {idx+2}: 'valor' inválido.")
                    valor_num = None  # type: ignore
                # Datas: se informadas e parse falhar, erro
                if row.get("dataInicio", None) not in (None, "") and dataInicio is None:
                    errors.append(f"Linha {idx+2}: 'dataInicio' inválida.")
                if row.get("dataFim", None) not in (None, "") and dataFim is None:
                    errors.append(f"Linha {idx+2}: 'dataFim' inválida.")

                # FK cliente
                if isinstance(cliente_id_int, int):
                    cli = db.query(Cliente).filter(Cliente.id == cliente_id_int).first()
                    if not cli:
                        errors.append(f"Linha {idx+2}: cliente_id {cliente_id_int} não existe.")

                rows_data.append(
                    {
                        "numero": numero,
                        "cliente_id": cliente_id_int,
                        "valor": valor_num,
                        "dataInicio": dataInicio,
                        "dataFim": dataFim,
                        "tipo": tipo,
                        "situacao": situacao,
                        "prazoPagamento": prazoPagamento,
                        "quantidadeParcelas": quantidadeParcelas,
                    }
                )

            if errors:
                # Rollback total e retornar erros (aceitar somente se 100% válido)
                db.rollback()
                # Limitar quantidade de erros na resposta
                preview = "; ".join(errors[:10])
                if len(errors) > 10:
                    preview += f" (+{len(errors)-10} erros)"
                raise HTTPException(status_code=400, detail=f"Erros de validação: {preview}")

            # Inserir (upsert por numero)
            for data in rows_data:
                existing = db.query(Contrato).filter(Contrato.numero == data["numero"]).first()
                if existing:
                    existing.cliente_id = data["cliente_id"]
                    existing.valor = data["valor"]
                    existing.dataInicio = data["dataInicio"]
                    existing.dataFim = data["dataFim"]
                    existing.tipo = data["tipo"]
                    existing.situacao = data["situacao"]
                    existing.prazoPagamento = data["prazoPagamento"]
                    existing.quantidadeParcelas = data["quantidadeParcelas"]
                    db.add(existing)
                else:
                    novo = Contrato(**data)
                    db.add(novo)
                new_count += 1
            db.commit()
            imported = new_count
        except HTTPException:
            raise
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"Falha ao importar contratos: {e}")

    return UploadResult(upload_id=up.id, filename=up.nome, entidade=up.entidade, records_imported=imported)

@app.get("/uploads")
@app.get("/api/uploads")
def list_uploads(entidade: Optional[str] = None, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    q = db.query(ArquivoImportado)
    if entidade:
        q = q.filter(ArquivoImportado.entidade == entidade)
    items = q.order_by(ArquivoImportado.id.desc()).all()
    return [
        {
            "id": it.id,
            "nome": it.nome,
            "entidade": it.entidade,
            "tamanho": it.tamanho,
            "criado_em": it.criado_em,
        }
        for it in items
    ]

@app.get("/uploads/{upload_id}/download")
@app.get("/api/uploads/{upload_id}/download")
def download_upload(upload_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    up = db.query(ArquivoImportado).filter(ArquivoImportado.id == upload_id).first()
    if not up:
        raise HTTPException(status_code=404, detail="Upload não encontrado")
    headers = {"Content-Disposition": f"attachment; filename={up.nome}"}
    return StreamingResponse(io.BytesIO(up.conteudo), headers=headers)

# ------------------------
# Modelos Excel (Templates)
# ------------------------
def _build_template_workbook(entidade: str):
    """Gera um workbook Excel em memória com colunas e uma linha de exemplo para a entidade informada.

    Retorna bytes do arquivo .xlsx.
    """
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as e:
        # Segurança: se openpyxl não estiver disponível, retorna 500 claro
        raise HTTPException(status_code=500, detail=f"Biblioteca openpyxl não disponível: {e}")

    entidade_norm = (entidade or "").strip().lower()

    # Definições por entidade (colunas e exemplo)
    templates: dict[str, dict[str, list]] = {
        "clientes": {
            "columns": ["id", "nome", "cnpj", "email", "contato", "endereco"],
            "example": ["", "Cliente Exemplo LTDA", "12.345.678/0001-99", "contato@exemplo.com", "Maria", "Rua A, 123"],
        },
        "valor_materiais": {
            # Observação: o import aceita variações (descricao, unidade, valor), mas sugerimos as colunas abaixo
            "columns": [
                "descricao_produto",
                "marca",
                "unidade_medida",
                "valor_unitario",
                "estoque_atual",
                "estoque_minimo",
                "data_ultima_entrada",
                "responsavel",
                "fornecedor",
                "localizacao",
                "observacoes",
            ],
            "example": [
                "Parafuso 1/2",
                "ACME",
                "un",
                2.5,
                100,
                10,
                "2025-01-10",
                "João",
                "Ferros XYZ",
                "Prateleira 3",
                "Lote novo",
            ],
        },
        "despesas": {
            "columns": ["id_cliente", "servico", "valor", "data", "categoria", "status", "observacoes"],
            "example": [1, "Manutenção preventiva", 1500.0, "2025-02-01", "Manutenção", "Pendente", "Troca de filtros"],
        },
        "fornecedores": {
            "columns": ["id", "nome", "cnpj", "contato"],
            "example": ["", "Fornecedor ABC", "23.456.789/0001-00", "Carlos"],
        },
        "contratos": {
            "columns": [
                "numero",
                "cliente_id",
                "valor",
                "dataInicio",
                "dataFim",
                "tipo",
                "situacao",
                "prazoPagamento",
                "quantidadeParcelas",
            ],
            "example": [
                "CT-2025-001",
                1,
                50000.0,
                "2025-03-01",
                "2025-09-30",
                "Obra",
                "Vigente",
                "30 dias",
                "6",
            ],
        },
        "orcamento_obra": {
            "columns": ["cliente_id", "etapa", "descricao", "unidade", "quantidade", "custo_unitario", "data"],
            "example": [1, "Alvenaria", "Tijolo cerâmico 8 furos", "m2", 120.0, 35.0, "2025-03-15"],
        },
        "resumo_mensal": {
            "columns": ["cliente_id", "mes", "ano", "total_despesas", "total_orcamento"],
            "example": [1, "Março", 2025, 12500.0, 20000.0],
        },
        "usuarios": {
            "columns": ["username", "password", "nome", "email", "nivel_acesso", "ativo", "grupo_id"],
            "example": ["usuario1", "senha123", "Usuário 1", "u1@dominio.com", "visualizacao", True, ""],
        },
    }

    cfg = templates.get(entidade_norm)
    if not cfg:
        # fallback: retorna apenas uma planilha vazia com instruções
        cfg = {"columns": ["coluna1", "coluna2"], "example": ["", ""]}

    wb = Workbook()
    ws = wb.active
    ws.title = "dados"
    # Cabeçalhos
    for j, col in enumerate(cfg["columns"], start=1):
        ws.cell(row=1, column=j, value=col)
    # Linha de exemplo
    for j, val in enumerate(cfg["example"], start=1):
        ws.cell(row=2, column=j, value=val)

    # Segunda aba com instruções
    readme = wb.create_sheet("LEIA-ME")
    instrucoes = [
        f"Modelo de importação para a entidade: {entidade_norm}",
        "Preencha a aba 'dados' com uma linha por registro.",
        "Não altere os nomes das colunas da primeira linha (cabeçalho).",
        "Colunas marcadas como opcionais podem ficar em branco.",
        "Datas devem estar no formato ISO (YYYY-MM-DD).",
        "Números devem usar ponto como separador decimal (ex.: 1234.56).",
        "Para Clientes: 'nome' é obrigatório; 'id' é opcional para atualizar registros existentes.",
        "Para Valor Materiais: utilize as colunas do template; variações como 'descricao' e 'unidade' também são aceitas no import.",
        "Para Despesas: use 'id_cliente' numérico existente no sistema.",
    ]
    for i, line in enumerate(instrucoes, start=1):
        readme.cell(row=i, column=1, value=line)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@app.get("/templates/{entidade}")
def download_template(entidade: str, current_user: Usuario = Depends(get_current_user)):
    data = _build_template_workbook(entidade)
    filename = f"template_{entidade.lower()}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

# Compat de rota específica usada no frontend para valor_materiais
@app.post("/upload_valor_materiais", response_model=UploadResult)
async def upload_valor_materiais(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    return await upload_entidade("valor_materiais", file, db, current_user)

# Alias explícito para importação de contratos sem colisão com static files
@app.post("/import/contratos", response_model=UploadResult)
async def import_contratos(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    return await upload_entidade("contratos", file, db, current_user)
@app.get("/me/permissoes")
def get_my_permissions(current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)):
    """Retorna permissões efetivas do usuário (via grupo)."""
    if not current_user.grupo_id:
        return {"grupo_id": None, "permissoes": []}
    # relacionamentos: PermissaoGrupo(grupo_id, permissao_id) → PermissaoSistema
    links = db.query(PermissaoGrupo).filter(PermissaoGrupo.grupo_id == current_user.grupo_id).all()
    ids = [lk.permissao_id for lk in links]
    if not ids:
        return {"grupo_id": current_user.grupo_id, "permissoes": []}
    perms = db.query(PermissaoSistema).filter(PermissaoSistema.id.in_(ids)).all()
    return {
        "grupo_id": current_user.grupo_id,
        "permissoes": [
            {"id": p.id, "nome": p.nome, "categoria": p.categoria, "ativo": p.ativo}
            for p in perms
        ]
    }

# Grupos de Usuários
@app.get("/grupos/", response_model=List[GrupoUsuarioSchema])
def list_grupos(db: Session = Depends(get_db), current_user: Usuario = Depends(require_admin)):
    return db.query(GrupoUsuario).all()

@app.post("/grupos/", response_model=GrupoUsuarioSchema)
def create_grupo(grupo: GrupoUsuarioCreateSchema, db: Session = Depends(get_db), current_user: Usuario = Depends(require_admin)):
    db_grupo = GrupoUsuario(
        nome=grupo.nome,
        descricao=grupo.descricao,
        status=grupo.status,
        motivo=grupo.motivo,
        valor_maximo_diario_financeiro=grupo.valor_maximo_diario_financeiro,
        preco_venda=grupo.preco_venda,
        plano_contas=grupo.plano_contas,
        valor_maximo_movimentacao=grupo.valor_maximo_movimentacao,
        valor_maximo_solicitacao_compra=grupo.valor_maximo_solicitacao_compra,
        valor_maximo_diario_solicitacao_compra=grupo.valor_maximo_diario_solicitacao_compra
    )
    # Adicionar o grupo primeiro e fazer o commit para obter o ID
    db.add(db_grupo)
    db.commit()
    db.refresh(db_grupo)
    
    # Processar permissões se fornecidas
    if grupo.permissoes:
        for permissao_id in grupo.permissoes:
            # Converter string para int se necessário
            perm_id = int(permissao_id) if isinstance(permissao_id, str) else permissao_id
            db_permissao_grupo = PermissaoGrupo(
                grupo_id=db_grupo.id,
                permissao_id=perm_id,
                ativo=True
            )
            db.add(db_permissao_grupo)
    
    # Processar lojas se fornecidas
    if grupo.lojas:
        for loja_id in grupo.lojas:
            db_loja_grupo = LojaGrupo(
                grupo_id=db_grupo.id,
                loja_id=loja_id,
                acesso_total=grupo.acesso_total_lojas
            )
            db.add(db_loja_grupo)
    # Processar clientes se fornecidos
    if grupo.clientes:
        for cliente_id in grupo.clientes:
            db_cliente_grupo = ClienteGrupo(
                grupo_id=db_grupo.id,
                cliente_id=cliente_id,
            )
            db.add(db_cliente_grupo)
            
    db.commit()
    db.refresh(db_grupo)
    return db_grupo

@app.get("/grupos/{grupo_id}", response_model=GrupoUsuarioSchema)
def get_grupo(grupo_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(require_admin)):
    grupo = db.query(GrupoUsuario).filter(GrupoUsuario.id == grupo_id).first()
    if not grupo:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")
    return grupo

@app.put("/grupos/{grupo_id}", response_model=GrupoUsuarioSchema)
def update_grupo(grupo_id: int, grupo: GrupoUsuarioCreateSchema, db: Session = Depends(get_db), current_user: Usuario = Depends(require_admin)):
    db_grupo = db.query(GrupoUsuario).filter(GrupoUsuario.id == grupo_id).first()
    if not db_grupo:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")
    
    # Atualizar campos do grupo
    update_data = grupo.dict(exclude_unset=True)
    for key, value in update_data.items():
        if key not in ["permissoes", "lojas"]:
            setattr(db_grupo, key, value)

    # Atualizar permissões
    if "permissoes" in update_data:
        # Limpar permissões existentes
        db.query(PermissaoGrupo).filter(PermissaoGrupo.grupo_id == grupo_id).delete()
        # Adicionar novas permissões
        for permissao_id in update_data["permissoes"]:
            perm_id = int(permissao_id) if isinstance(permissao_id, str) else permissao_id
            db_permissao_grupo = PermissaoGrupo(grupo_id=grupo_id, permissao_id=perm_id, ativo=True)
            db.add(db_permissao_grupo)

    # Atualizar lojas
    if "lojas" in update_data:
        # Limpar lojas existentes
        db.query(LojaGrupo).filter(LojaGrupo.grupo_id == grupo_id).delete()
        # Adicionar novas lojas
        for loja_id in update_data["lojas"]:
            db_loja_grupo = LojaGrupo(grupo_id=grupo_id, loja_id=loja_id, acesso_total=grupo.acesso_total_lojas)
            db.add(db_loja_grupo)
    # Atualizar clientes
    if "clientes" in update_data:
        db.query(ClienteGrupo).filter(ClienteGrupo.grupo_id == grupo_id).delete()
        for cliente_id in update_data["clientes"]:
            db_cliente_grupo = ClienteGrupo(grupo_id=grupo_id, cliente_id=cliente_id)
            db.add(db_cliente_grupo)

    db.commit()
    db.refresh(db_grupo)
    return db_grupo

@app.delete("/grupos/{grupo_id}")
def delete_grupo(grupo_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(require_admin)):
    db_grupo = db.query(GrupoUsuario).filter(GrupoUsuario.id == grupo_id).first()
    if not db_grupo:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")
    db.delete(db_grupo)
    db.commit()
    return {"ok": True}

# Permissões
@app.get("/permissoes/", response_model=List[PermissaoSistemaSchema])
def list_permissoes(db: Session = Depends(get_db), current_user: Usuario = Depends(require_admin)):
    return db.query(PermissaoSistema).all()

# Permissões de Grupo
@app.get("/grupos/{grupo_id}/permissoes/")
def get_grupo_permissoes(grupo_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(require_admin)):
    # Verificar se o grupo existe
    db_grupo = db.query(GrupoUsuario).filter(GrupoUsuario.id == grupo_id).first()
    if not db_grupo:
        raise HTTPException(status_code=404, detail="Grupo não encontrado")
    
    # Buscar permissões do grupo
    permissoes_grupo = db.query(PermissaoGrupo).filter(PermissaoGrupo.grupo_id == grupo_id).all()
    permissoes_ids = [pg.permissao_id for pg in permissoes_grupo]
    
    # Buscar detalhes das permissões
    permissoes = db.query(PermissaoSistema).filter(PermissaoSistema.id.in_(permissoes_ids)).all()
    
    return {
        "grupo_id": grupo_id,
        "grupo_nome": db_grupo.nome,
        "permissoes": [
            {
                "id": perm.id,
                "nome": perm.nome,
                "descricao": perm.descricao
            } for perm in permissoes
        ]
    }

# Lojas
@app.get("/lojas/", response_model=List[LojaSchema])
def list_lojas(db: Session = Depends(get_db)):
    return db.query(Loja).all()

@app.post("/lojas/", response_model=LojaSchema)
def create_loja(loja: LojaCreateSchema, db: Session = Depends(get_db)):
    db_loja = Loja(**loja.dict())
    db.add(db_loja)
    db.commit()
    db.refresh(db_loja)
    return db_loja

# Finalizando o arquivo main.py com endpoints de clientes e outras entidades
@app.get("/clientes/", response_model=List[ClienteSchema])
def listar_clientes(db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1201, "read"))):
    allowed = _get_allowed_client_ids(db, current_user)
    q = db.query(Cliente)
    if allowed is not None and len(allowed) > 0:
        q = q.filter(Cliente.id.in_(allowed))
    elif allowed is not None and len(allowed) == 0:
        return []
    return q.all()

@app.post("/clientes/", response_model=ClienteSchema)
def criar_cliente(cliente: ClienteCreateSchema, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1201, "create"))):
    # Admin/Willians podem criar qualquer cliente; demais, apenas se pertencente ao seu escopo (após criação, será necessário associar no grupo via UI)
    novo = Cliente(**cliente.dict())
    db.add(novo)
    db.commit()
    db.refresh(novo)
    return novo

# Despesas
@app.get("/despesas/", response_model=List[DespesaSchema])
def listar_despesas(db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1601, "read"))):
    return db.query(Despesa).all()

@app.get("/despesas", response_model=List[DespesaSchema])
def listar_despesas_alt(db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1601, "read"))):
    return db.query(Despesa).all()

@app.post("/despesas/", response_model=DespesaSchema)
def criar_despesa(despesa: DespesaSchema, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1601, "create"))):
    novo = Despesa(**despesa.dict(exclude_unset=True))
    db.add(novo)
    db.commit()
    db.refresh(novo)
    return novo

@app.put("/despesas/{despesa_id}", response_model=DespesaSchema)
def atualizar_despesa(despesa_id: int, despesa: DespesaSchema, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1601, "update"))):
    db_despesa = db.query(Despesa).filter(Despesa.id == despesa_id).first()
    if not db_despesa:
        raise HTTPException(status_code=404, detail="Despesa não encontrada")
    
    for key, value in despesa.dict(exclude_unset=True).items():
        if key != 'id':
            setattr(db_despesa, key, value)
    
    db.commit()
    db.refresh(db_despesa)
    return db_despesa

@app.delete("/despesas/{despesa_id}")
def deletar_despesa(despesa_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1601, "delete"))):
    db_despesa = db.query(Despesa).filter(Despesa.id == despesa_id).first()
    if not db_despesa:
        raise HTTPException(status_code=404, detail="Despesa não encontrada")
    db.delete(db_despesa)
    db.commit()
    return {"ok": True}

# Resumo Mensal
@app.get("/resumo_mensal/", response_model=List[ResumoMensalSchema])
def listar_resumo_mensal(db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1801, "read"))):
    return db.query(ResumoMensal).all()

@app.post("/resumo_mensal/", response_model=ResumoMensalSchema)
def criar_resumo_mensal(resumo: ResumoMensalSchema, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1801, "create"))):
    novo = ResumoMensal(**resumo.dict(exclude_unset=True))
    db.add(novo)
    db.commit()
    db.refresh(novo)
    return novo

@app.put("/resumo_mensal/{resumo_id}", response_model=ResumoMensalSchema)
def atualizar_resumo_mensal(resumo_id: int, resumo: ResumoMensalSchema, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1801, "update"))):
    db_resumo = db.query(ResumoMensal).filter(ResumoMensal.id == resumo_id).first()
    if not db_resumo:
        raise HTTPException(status_code=404, detail="Resumo mensal não encontrado")
    
    for key, value in resumo.dict(exclude_unset=True).items():
        if key != 'id':
            setattr(db_resumo, key, value)
    
    db.commit()
    db.refresh(db_resumo)
    return db_resumo

@app.delete("/resumo_mensal/{resumo_id}")
def deletar_resumo_mensal(resumo_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1801, "delete"))):
    db_resumo = db.query(ResumoMensal).filter(ResumoMensal.id == resumo_id).first()
    if not db_resumo:
        raise HTTPException(status_code=404, detail="Resumo mensal não encontrado")
    db.delete(db_resumo)
    db.commit()
    return {"ok": True}

# Fornecedores
@app.get("/fornecedores/", response_model=List[FornecedorSchema])
def listar_fornecedores(db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1301, "read"))):
    return db.query(Fornecedor).all()

@app.post("/fornecedores/", response_model=FornecedorSchema)
def criar_fornecedor(fornecedor: FornecedorSchema, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1301, "create"))):
    novo = Fornecedor(**fornecedor.dict(exclude_unset=True))
    db.add(novo)
    db.commit()
    db.refresh(novo)
    return novo

# Orçamento de Obra
@app.get("/orcamento_obra/", response_model=List[OrcamentoObraSchema])
def listar_orcamento_obra(db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1501, "read"))):
    return db.query(OrcamentoObra).all()

@app.post("/orcamento_obra/", response_model=OrcamentoObraSchema)
def criar_orcamento_obra(orcamento: OrcamentoObraSchema, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1501, "create"))):
    novo = OrcamentoObra(**orcamento.dict(exclude_unset=True))
    db.add(novo)
    db.commit()
    db.refresh(novo)
    return novo

# Contratos
@app.get("/contratos/", response_model=List[ContratoSchema])
def listar_contratos(db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1401, "read"))):
    return db.query(Contrato).all()

@app.post("/contratos/", response_model=ContratoSchema)
def criar_contrato(contrato: ContratoSchema, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1401, "create"))):
    novo = Contrato(**contrato.dict(exclude_unset=True))
    db.add(novo)
    db.commit()
    db.refresh(novo)
    return novo

# Valor de Materiais
@app.get("/valor_materiais/", response_model=List[ValorMaterialSchema])
def listar_valor_materiais(db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1701, "read"))):
    return db.query(ValorMaterial).all()

@app.post("/valor_materiais/", response_model=ValorMaterialSchema)
def criar_valor_material(material: ValorMaterialSchema, db: Session = Depends(get_db), current_user: Usuario = Depends(_permission_required(1701, "create"))):
    novo = ValorMaterial(**material.dict(exclude_unset=True))
    db.add(novo)
    db.commit()
    db.refresh(novo)
    return novo

# Testes de Loja
@app.get("/testes-loja/", response_model=List[TesteLojaSchema])
def listar_testes_loja(db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    allowed = _get_allowed_client_ids(db, current_user)
    q = db.query(TesteLoja)
    if allowed is not None and len(allowed) > 0:
        q = q.filter(TesteLoja.cliente_id.in_(allowed))
    elif allowed is not None and len(allowed) == 0:
        return []
    return q.all()

@app.get("/testes-loja/{teste_id}", response_model=TesteLojaSchema)
def obter_teste_loja(teste_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    teste = db.query(TesteLoja).filter(TesteLoja.id == teste_id).first()
    if not teste:
        raise HTTPException(status_code=404, detail="Teste não encontrado")
    allowed = _get_allowed_client_ids(db, current_user)
    if allowed is not None and len(allowed) > 0 and teste.cliente_id not in allowed:
        raise HTTPException(status_code=403, detail="Sem acesso a este cliente")
    return teste

@app.post("/testes-loja/", response_model=TesteLojaSchema)
async def criar_teste_loja(
    data_teste: str = Form(...),
    cliente_id: int = Form(...),
    horario: str = Form(...),
    status: str = Form(...),
    observacao: str = Form(None),
    foto: UploadFile = File(None),
    video: UploadFile = File(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    # Verificar permissão por cliente_id
    allowed = _get_allowed_client_ids(db, current_user)
    if allowed is not None and cliente_id not in allowed:
        raise HTTPException(status_code=403, detail="Sem acesso a este cliente")
    # Criar teste primeiro
    novo_teste = TesteLoja(
        data_teste=datetime.datetime.strptime(data_teste, "%Y-%m-%d").date(),
        cliente_id=cliente_id,
        horario=datetime.datetime.strptime(horario, "%H:%M").time(),
        status=status,
        observacao=observacao
    )
    db.add(novo_teste)
    db.commit()
    db.refresh(novo_teste)
    
    # Salvar arquivos se fornecidos
    if foto:
        file_extension = os.path.splitext(foto.filename)[1]
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_path = os.path.join(TESTES_LOJA_DIR, unique_filename)
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(foto.file, buffer)
        
        novo_teste.foto = unique_filename
    
    if video:
        file_extension = os.path.splitext(video.filename)[1]
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_path = os.path.join(TESTES_LOJA_DIR, unique_filename)
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(video.file, buffer)
        
        novo_teste.video = unique_filename
    
    db.commit()
    db.refresh(novo_teste)
    return novo_teste

@app.put("/testes-loja/{teste_id}", response_model=TesteLojaSchema)
async def atualizar_teste_loja(
    teste_id: int,
    data_teste: str = Form(...),
    cliente_id: int = Form(...),
    horario: str = Form(...),
    status: str = Form(...),
    observacao: str = Form(None),
    foto: UploadFile = File(None),
    video: UploadFile = File(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    teste_db = db.query(TesteLoja).filter(TesteLoja.id == teste_id).first()
    if not teste_db:
        raise HTTPException(status_code=404, detail="Teste não encontrado")
    
    # Checar acesso ao novo cliente_id
    allowed = _get_allowed_client_ids(db, current_user)
    if allowed is not None and cliente_id not in allowed:
        raise HTTPException(status_code=403, detail="Sem acesso a este cliente")
    # Atualizar dados básicos
    teste_db.data_teste = datetime.datetime.strptime(data_teste, "%Y-%m-%d").date()
    teste_db.cliente_id = cliente_id
    teste_db.horario = datetime.datetime.strptime(horario, "%H:%M").time()
    teste_db.status = status
    teste_db.observacao = observacao
    
    # Salvar novos arquivos se fornecidos
    if foto and foto.filename:
        # Remover arquivo antigo se existir
        if teste_db.foto:
            old_file_path = os.path.join(TESTES_LOJA_DIR, teste_db.foto)
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
        
        # Salvar novo arquivo
        file_extension = os.path.splitext(foto.filename)[1]
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_path = os.path.join(TESTES_LOJA_DIR, unique_filename)
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(foto.file, buffer)
        
        teste_db.foto = unique_filename
    
    if video and video.filename:
        # Remover arquivo antigo se existir
        if teste_db.video:
            old_file_path = os.path.join(TESTES_LOJA_DIR, teste_db.video)
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
        
        # Salvar novo arquivo
        file_extension = os.path.splitext(video.filename)[1]
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_path = os.path.join(TESTES_LOJA_DIR, unique_filename)
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(video.file, buffer)
        
        teste_db.video = unique_filename
    
    db.commit()
    db.refresh(teste_db)
    return teste_db

@app.delete("/testes-loja/{teste_id}")
def deletar_teste_loja(teste_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    teste_db = db.query(TesteLoja).filter(TesteLoja.id == teste_id).first()
    if not teste_db:
        raise HTTPException(status_code=404, detail="Teste não encontrado")
    # Checar escopo de cliente
    allowed = _get_allowed_client_ids(db, current_user)
    if allowed is not None and teste_db.cliente_id not in allowed:
        raise HTTPException(status_code=403, detail="Sem acesso a este cliente")
    
    db.delete(teste_db)
    db.commit()
    return {"message": "Teste deletado com sucesso"}

# Testes de Ar Condicionado
@app.get("/testes-ar-condicionado/", response_model=List[TesteArCondicionadoSchema])
def listar_testes_ar_condicionado(db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    allowed = _get_allowed_client_ids(db, current_user)
    q = db.query(TesteArCondicionado)
    if allowed is not None and len(allowed) > 0:
        q = q.filter(TesteArCondicionado.cliente_id.in_(allowed))
    elif allowed is not None and len(allowed) == 0:
        return []
    return q.all()

@app.post("/testes-ar-condicionado/", response_model=TesteArCondicionadoSchema)
async def criar_teste_ar_condicionado(
    data_teste: str = Form(...),
    cliente_id: int = Form(...),
    horario: str = Form(...),
    status: str = Form(...),
    observacao: str = Form(None),
    foto: UploadFile = File(None),
    video: UploadFile = File(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    # Verificar permissão por cliente_id
    allowed = _get_allowed_client_ids(db, current_user)
    if allowed is not None and cliente_id not in allowed:
        raise HTTPException(status_code=403, detail="Sem acesso a este cliente")
    # Criar teste primeiro
    novo_teste = TesteArCondicionado(
        data_teste=datetime.datetime.strptime(data_teste, "%Y-%m-%d").date(),
        cliente_id=cliente_id,
        horario=datetime.datetime.strptime(horario, "%H:%M").time(),
        status=status,
        observacao=observacao
    )
    db.add(novo_teste)
    db.commit()
    db.refresh(novo_teste)
    
    # Salvar arquivos se fornecidos
    if foto:
        file_extension = os.path.splitext(foto.filename)[1]
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_path = os.path.join(TESTES_AR_CONDICIONADO_DIR, unique_filename)
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(foto.file, buffer)
        
        novo_teste.foto = unique_filename
    
    if video:
        file_extension = os.path.splitext(video.filename)[1]
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_path = os.path.join(TESTES_AR_CONDICIONADO_DIR, unique_filename)
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(video.file, buffer)
        
        novo_teste.video = unique_filename
    
    db.commit()
    db.refresh(novo_teste)
    return novo_teste

@app.put("/testes-ar-condicionado/{teste_id}", response_model=TesteArCondicionadoSchema)
async def atualizar_teste_ar_condicionado(
    teste_id: int,
    data_teste: str = Form(...),
    cliente_id: int = Form(...),
    horario: str = Form(...),
    status: str = Form(...),
    observacao: str = Form(None),
    foto: UploadFile = File(None),
    video: UploadFile = File(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    teste_db = db.query(TesteArCondicionado).filter(TesteArCondicionado.id == teste_id).first()
    if not teste_db:
        raise HTTPException(status_code=404, detail="Teste não encontrado")
    
    # Checar acesso ao novo cliente_id
    allowed = _get_allowed_client_ids(db, current_user)
    if allowed is not None and cliente_id not in allowed:
        raise HTTPException(status_code=403, detail="Sem acesso a este cliente")
    # Atualizar dados básicos
    teste_db.data_teste = datetime.datetime.strptime(data_teste, "%Y-%m-%d").date()
    teste_db.cliente_id = cliente_id
    teste_db.horario = datetime.datetime.strptime(horario, "%H:%M").time()
    teste_db.status = status
    teste_db.observacao = observacao
    
    # Salvar novos arquivos se fornecidos
    if foto and foto.filename:
        # Remover arquivo antigo se existir
        if teste_db.foto:
            old_file_path = os.path.join(TESTES_AR_CONDICIONADO_DIR, teste_db.foto)
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
        
        # Salvar novo arquivo
        file_extension = os.path.splitext(foto.filename)[1]
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_path = os.path.join(TESTES_AR_CONDICIONADO_DIR, unique_filename)
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(foto.file, buffer)
        
        teste_db.foto = unique_filename
    
    if video and video.filename:
        # Remover arquivo antigo se existir
        if teste_db.video:
            old_file_path = os.path.join(TESTES_AR_CONDICIONADO_DIR, teste_db.video)
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
        
        # Salvar novo arquivo
        file_extension = os.path.splitext(video.filename)[1]
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_path = os.path.join(TESTES_AR_CONDICIONADO_DIR, unique_filename)
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(video.file, buffer)
        
        teste_db.video = unique_filename
    
    db.commit()
    db.refresh(teste_db)
    return teste_db

@app.delete("/testes-ar-condicionado/{teste_id}")
def deletar_teste_ar_condicionado(teste_id: int, db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    teste_db = db.query(TesteArCondicionado).filter(TesteArCondicionado.id == teste_id).first()
    if not teste_db:
        raise HTTPException(status_code=404, detail="Teste não encontrado")
    # Checar escopo de cliente
    allowed = _get_allowed_client_ids(db, current_user)
    if allowed is not None and teste_db.cliente_id not in allowed:
        raise HTTPException(status_code=403, detail="Sem acesso a este cliente")
    
    db.delete(teste_db)
    db.commit()
    return {"message": "Teste deletado com sucesso"}

# Rotas de Upload para Testes
@app.post("/upload/testes-loja/foto/{teste_id}")
async def upload_foto_teste_loja(teste_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    # Verificar se o teste existe
    teste = db.query(TesteLoja).filter(TesteLoja.id == teste_id).first()
    if not teste:
        raise HTTPException(status_code=404, detail="Teste não encontrado")
    
    # Verificar escopo de cliente
    allowed = _get_allowed_client_ids(db, current_user)
    if allowed is not None and teste.cliente_id not in allowed:
        raise HTTPException(status_code=403, detail="Sem acesso a este cliente")

    # Gerar nome único para o arquivo
    file_extension = os.path.splitext(file.filename)[1]
    unique_filename = f"{uuid.uuid4()}{file_extension}"
    file_path = os.path.join(TESTES_LOJA_DIR, unique_filename)
    
    # Salvar arquivo
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Atualizar teste no banco
    teste.foto = unique_filename
    db.commit()
    
    return {"filename": unique_filename, "message": "Foto enviada com sucesso"}

@app.post("/upload/testes-loja/video/{teste_id}")
async def upload_video_teste_loja(teste_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    # Verificar se o teste existe
    teste = db.query(TesteLoja).filter(TesteLoja.id == teste_id).first()
    if not teste:
        raise HTTPException(status_code=404, detail="Teste não encontrado")
    
    # Verificar escopo de cliente
    allowed = _get_allowed_client_ids(db, current_user)
    if allowed is not None and teste.cliente_id not in allowed:
        raise HTTPException(status_code=403, detail="Sem acesso a este cliente")

    # Gerar nome único para o arquivo
    file_extension = os.path.splitext(file.filename)[1]
    unique_filename = f"{uuid.uuid4()}{file_extension}"
    file_path = os.path.join(TESTES_LOJA_DIR, unique_filename)
    
    # Salvar arquivo
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Atualizar teste no banco
    teste.video = unique_filename
    db.commit()
    
    return {"filename": unique_filename, "message": "Vídeo enviado com sucesso"}

@app.post("/upload/testes-ar-condicionado/foto/{teste_id}")
async def upload_foto_teste_ar_condicionado(teste_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    # Verificar se o teste existe
    teste = db.query(TesteArCondicionado).filter(TesteArCondicionado.id == teste_id).first()
    if not teste:
        raise HTTPException(status_code=404, detail="Teste não encontrado")
    
    # Verificar escopo de cliente
    allowed = _get_allowed_client_ids(db, current_user)
    if allowed is not None and teste.cliente_id not in allowed:
        raise HTTPException(status_code=403, detail="Sem acesso a este cliente")

    # Gerar nome único para o arquivo
    file_extension = os.path.splitext(file.filename)[1]
    unique_filename = f"{uuid.uuid4()}{file_extension}"
    file_path = os.path.join(TESTES_AR_CONDICIONADO_DIR, unique_filename)
    
    # Salvar arquivo
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Atualizar teste no banco
    teste.foto = unique_filename
    db.commit()
    
    return {"filename": unique_filename, "message": "Foto enviada com sucesso"}

@app.post("/upload/testes-ar-condicionado/video/{teste_id}")
async def upload_video_teste_ar_condicionado(teste_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), current_user: Usuario = Depends(get_current_user)):
    # Verificar se o teste existe
    teste = db.query(TesteArCondicionado).filter(TesteArCondicionado.id == teste_id).first()
    if not teste:
        raise HTTPException(status_code=404, detail="Teste não encontrado")
    
    # Verificar escopo de cliente
    allowed = _get_allowed_client_ids(db, current_user)
    if allowed is not None and teste.cliente_id not in allowed:
        raise HTTPException(status_code=403, detail="Sem acesso a este cliente")

    # Gerar nome único para o arquivo
    file_extension = os.path.splitext(file.filename)[1]
    unique_filename = f"{uuid.uuid4()}{file_extension}"
    file_path = os.path.join(TESTES_AR_CONDICIONADO_DIR, unique_filename)
    
    # Salvar arquivo
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Atualizar teste no banco
    teste.video = unique_filename
    db.commit()
    
    return {"filename": unique_filename, "message": "Vídeo enviado com sucesso"}

# Endpoint básico para testar o servidor
@app.get("/")
def root():
    return {"message": "API de Gestão de Obras funcionando!"}

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
